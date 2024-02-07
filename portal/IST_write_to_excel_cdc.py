from django.shortcuts import render
from django.http import FileResponse
import os
import datetime
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import pyplot
import seaborn as sns
import matplotlib as mpl
import seaborn as sns  
from scipy.stats import norm
from scipy.stats import ttest_ind
import statistics
from IPython.display import Image
import io
import requests
from openpyxl import load_workbook
import os
import glob
from openpyxl import load_workbook
import shutil
import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime, timedelta
import pyodbc
import urllib
import smtplib
from email.mime.text import MIMEText
import warnings
import time
import random
from django.contrib import messages
from django.core.mail import message
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render,redirect
from django.contrib.auth import authenticate,login,logout
from django.core.mail import send_mail
from django.http import HttpResponse
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from sqlalchemy.pool import QueuePool
import traceback
from datetime import datetime
from datetime import timedelta
from datetime import datetime, timedelta
# Additional necessary imports...

# Utility function to get the last Sunday
def get_last_sunday(date):
    return date - datetime.timedelta(days=date.weekday() + 1)

# Utility function to create and copy the Excel template
def IST_create_and_copy_excel_template(template_path, destination_folder_base):
    today = datetime.today()
    current_year = today.year
    current_month = today.strftime('%B')
    # Get the current date and time
    current_date = datetime.now()
    # Check if today is Sunday
    if current_date.weekday() == 6:  # In Python, Monday is 0, so Sunday is 6
        recent_sunday = current_date
    else:
        # Calculate the most recent Sunday
        recent_sunday = current_date - timedelta(days=current_date.weekday() + 1)
    # Format the date in a readable format (e.g., YYYY-MM-DD)
    last_sunday_date = recent_sunday.strftime("%Y-%m-%d")

    year_folder = os.path.join(destination_folder_base, str(current_year))
    month_folder = os.path.join(year_folder, current_month)
    os.makedirs(month_folder, exist_ok=True)

    new_file_name = f'National IST Rider and Vehicle Weekly Report_{last_sunday_date}.xlsx'
    new_file_path = os.path.join(month_folder, new_file_name)

    shutil.copy2(template_path, new_file_path)

    return new_file_path

# Function to write data to the Excel file
def IST_write_data_to_excel(file_path, df_rider, df_relief,df_driver):
    # Open the Excel file
    workbook = load_workbook(file_path)
    #sheet = workbook.active
    # sheet_rider = workbook['Rider Weekly'] 
    # sheet_relief = workbook['Relief Rider Weekly'] 
    # sheet_driver = workbook['Provincial Driver Weekly'] 
    # Access sheets by index
    sheet_rider = workbook.worksheets[0]  # First sheet
    sheet_relief = workbook.worksheets[1]  # Second sheet
    sheet_driver = workbook.worksheets[2]  # Third sheet


    
    # Column mappings (DataFrame column name to Excel cell column) ---Received VL.sql
    column_mappings = {
       'Name_of_Rider_': 'A', 
       'Bike_Registration_Number': 'B', 
       'Province_': 'C', 
       'District_': 'D',
       'Type_of_PEPFAR_Support': 'E',
       'vl_plasma_sam': 'F', 
       'vl_dbs_sam': 'G', 
       'eid_sam': 'H',
       'eid_dbs': 'I',
       'sputum_sam': 'J', 
       'Sputum_Culture_DR_NTBRL':'K', 
       'HPV':'L', 
       'other_sam':'M',
       'vl_plasma_res':'N', 
       'vl_dbs_res':'O', 
       'eid_res':'P', 
       'eid_dbs_res':'Q', 
       'sputum_res':'R',
       'Sputum_Culture_DR_NTBRL_res':'S', 
       'HPV_res':'T', 
       'other_res':'U',
       'Fuel_allocated_to_rider_per_week':'V', 
       'Fuel_used_by_rider_per_week':'W',
       'Distance_travelled_by_rider_per_week':'X',
       'Number_of_days_bike_was_functional':'Y',
       'Number_of__Scheduled_Visits_to_Clinic_per__Week':'Z',
       'Number_of_Visits_to_Clinic_per_week':'AA', 
       'Bike_breakdown_':'AB',
       'Bike_on_routine_service_and_mainte0nce':'AC', 
       'Bike_had_no_fuel':'AD',
       'Rider_on_Sick_Leave':'AE', 
       'Rider_on_Leave':'AF', 
       'Inclement_weather':'AG',
       'Accident_damaged_bike_vehicle':'AH', 
       'Clinical_IPs_related_issues':'AI',
       'Other_Reasons__Specify_':'AJ', 
       'Mitigation_measures_':'AK', 
       'Comments':'AL'
    }

        # Starting row in the Excel sheet
    start_row = 3

    # Iterate through each row in the DataFrame and write to Rider Weekly
    for index, row in df_rider.iterrows():
        excel_row = start_row + index
        for col_name, excel_col in column_mappings.items():
            sheet_rider[f'{excel_col}{excel_row}'] = row[col_name]

    # Iterate through each row in the df_referred DataFrame and write to Relief Rider Weekly
    for index, row in df_relief.iterrows():
        excel_row = start_row + index
        for col_name, excel_col in column_mappings.items():
            sheet_relief[f'{excel_col}{excel_row}'] = row[col_name]

                

                
    #-------------------Results dispatched ---------------------results dispatch.sql
    Driver_mappings = {
        
       'Driver_Sample_Status': 'A',
       'Name_of_Rider_': 'B',
       'Bike_Registration_Number': 'C',
       'Province_': 'D', 
       'District_': 'E', 
       'vl_plasma_sam':'F',
       'eid_sam': 'H',
       'vl_dbs_sam':'G',
       'eid_dbs':'I', 
       'sputum_sam':'J', 
       'Sputum_Culture_DR_NTBRL':'K', 
       'HPV':'L', 
       'other_sam' :'M',
       'vl_plasma_res': 'N', 
       'eid_res':'O', 
       'vl_dbs_res':'P', 
       'eid_dbs_res':'Q',
       'sputum_res':'R',
       'Sputum_Culture_DR_NTBRL_res':'S',
       'HPV_res':'T', 
       'other_res':'U',
       'Fuel_allocated_to_rider_per_week': 'V',
       'Fuel_used_by_rider_per_week':'W',
       'Distance_travelled_by_rider_per_week':'X',
       'Number_of_days_bike_was_functional':'Y',
       'Number_of__Scheduled_Visits_to_Clinic_per__Week':'Z',
       'Number_of_Visits_to_Clinic_per_week':'AA', 
       'Bike_breakdown_':'AB',
       'Bike_on_routine_service_and_mainte0nce':'AC',
       'Bike_had_no_fuel':'AD',
       'Rider_on_Sick_Leave':'AE', 
       'Rider_on_Leave':'AF',
       'Inclement_weather':'AG',
       'Accident_damaged_bike_vehicle':'AH',
       'Clinical_IPs_related_issues':'AI',
       'Other_Reasons__Specify_':'AJ',
       'Mitigation_measures_':'AK', 
       'Comments':'AL'
        
              }

    for index, row in df_driver.iterrows():
        excel_row = start_row + index
        for col_name, excel_col in Driver_mappings.items():
            sheet_driver[f'{excel_col}{excel_row}'] = row[col_name]



    # Save the workbook
    workbook.save(file_path)





