from django.shortcuts import render
from django.http import FileResponse
import os
import datetime
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import pyplot
import seaborn as sns
import matplotlib as mpl
import seaborn as sns  
from scipy.stats import norm
from scipy.stats import ttest_ind
import statistics
from IPython.display import Image
import io
import requests
from openpyxl import load_workbook
import os
import glob
from openpyxl import load_workbook
import shutil
import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime, timedelta
import pyodbc
import urllib
import smtplib
from email.mime.text import MIMEText
import warnings
import time
import random
from django.contrib import messages
from django.core.mail import message
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render,redirect
from django.contrib.auth import authenticate,login,logout
from django.core.mail import send_mail
from django.http import HttpResponse
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from sqlalchemy.pool import QueuePool
import traceback
from datetime import datetime
from datetime import timedelta
from datetime import datetime, timedelta
# Additional necessary imports...

# Utility function to get the last Sunday
def get_last_sunday(date):
    return date - datetime.timedelta(days=date.weekday() + 1)

# Utility function to create and copy the Excel template
def create_and_copy_excel_template(template_path, destination_folder_base):
    today = datetime.today()
    current_year = today.year
    current_month = today.strftime('%B')
    # Get the current date and time
    current_date = datetime.now()
    # Check if today is Sunday
    if current_date.weekday() == 6:  # In Python, Monday is 0, so Sunday is 6
        recent_sunday = current_date
    else:
        # Calculate the most recent Sunday
        recent_sunday = current_date - timedelta(days=current_date.weekday() + 1)
    # Format the date in a readable format (e.g., YYYY-MM-DD)
    last_sunday_date = recent_sunday.strftime("%Y-%m-%d")

    year_folder = os.path.join(destination_folder_base, str(current_year))
    month_folder = os.path.join(year_folder, current_month)
    os.makedirs(month_folder, exist_ok=True)

    new_file_name = f'BRTI VL-EID Weekly Statistics Tool_{last_sunday_date}.xlsx'
    new_file_path = os.path.join(month_folder, new_file_name)

    shutil.copy2(template_path, new_file_path)

    return new_file_path

# Function to write data to the Excel file
def write_data_to_excel(file_path, df, df_eid_rec, df_downtime, df_ref_reasons, df_referred, df_dispatch, df_VL_run, df_EID_run, df_vl_run_reasons, df_eid_run_reasons,df_power_outage):
    # Open the Excel file
    workbook = load_workbook(file_path)
    #sheet = workbook.active
    sheet = workbook['CDC']  # Assuming 'CDC' is your sheet name

    
    # Column mappings (DataFrame column name to Excel cell column) ---Received VL.sql
    column_mappings = {
        'Carryover_Samples_in_the_lab_Plasma_VL': 'D',
        'Carryover_Samples_in_the_lab_DBS_VL': 'E',
        'Samples_in_backlog_Intra_lab_TAT_7_days_Plasma_VL': 'F',
        'Samples_in_backlog_Intra_lab_TAT_7_days_DBS_VL': 'G',
        'Total_samples_received_Plasma_VL': 'H',
        'Total_samples_received_DBS_VL': 'I',
        'Num_Rejected_Samples_Plasma_VL': 'J',
        'Num_Rejected_Samples_DBS_VL': 'K',
        'LIMS_logged_during_week_of_arrival_Plasma_VL': 'N',
        'LIMS_logged_during_week_of_arrival_DBS_VL': 'O',
        'LIMS_hub_logged_prior_to_arrival_Plasma_VL': 'P',
        'LIMS_hub_logged_prior_to_arrival_DBS_VL': 'Q',
        'LLIMs_Backlog_yetTObeEntered_Plasma_VL': 'R',
        'LIMs_Backlog_yetTObeEntered_DBS_VL': 'S',
        
        'REJECTED_too_old_to_test_Plasma_VL' : 'AJ', 
        'REJECTED_too_old_to_test_DBS_VL' :'AK',
        'REJECTED_Quality_issue_Plasma_VL' : 'AL', 
        'REJECTED_Quality_issue_DBS_VL' : 'AM',
        'REJECTED_Quantity_insuff_Plasma_VL' : 'AN',
        'REJECTED_Quantity_insuff_DBS_VL' : 'AO',
       'REJECTED_Quanti_Quali_intransit_compromised_Plasma_VL' : 'AP',
       'REJECTED_Quanti_Quali_intransit_compromised_DBS_VL' : 'AQ',
       'REJECTED_Patient_SampleINFO_Plasma_VL' : 'AR',
       'REJECTED_Patient_SampleINFO_DBS_VL' : 'AS',
       'REJECTED_Missing_requestForm_Plasma_VL' : 'AT',
       'REJECTED_Missing_requestForm_DBS_VL' : 'AU',
       'REJECTED_Sample_Missing_Plasma_VL' : 'AV', 
        'REJECTED_Sample_Missing_DBS_VL' :'AW',
       'REJECTED_other_reasons_Plasma_VL' : 'AX',
        'REJECTED_other_reasons_DBS_VL' : 'AY',
       'Comments' : 'BJ'
    }

    # Starting row in the Excel sheet
    start_row = 13

    # Iterate through each row in the DataFrame
    for index, row in df.iterrows():
        # Find the corresponding row in the Excel sheet
        lab_name = row['Lab']

        excel_row = None
        for row_num in range(start_row, start_row + 14):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == lab_name:
                excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in column_mappings.items():          
                sheet[f'{excel_col}{excel_row}'] = row[col_name]
                
                

                
    #------------------ referred module------------ reffereed.sql
    referred_mappings = {
        
           'Hrs_of_Fxnty': 'T',
           'Plasma_Samples_reffered_Out' : 'U',
           'DBS_Samples_reffered_Out' : 'V', 
           'Plasma_refered_to_lab' : 'W',
           'DBS_refered_to_lab' : 'X',
           'Plasma_Referred_Sample_Received': 'Y',
           'DBS_Referred_Sample_Received': 'Z',
           'Plasma_Referred_From_lab': 'AA',
           'DBS_Referred_From_lab': 'AB'  
        
    }

    # Iterate through each row in the DataFrame
    for index, row in df_referred.iterrows():
        # Find the corresponding row in the Excel sheet
        reflab_name = row['name_of_lab']

        ref_excel_row = None
        for row_num in range(start_row, start_row + 14):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == reflab_name:
                ref_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if ref_excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in referred_mappings.items():          
                sheet[f'{excel_col}{ref_excel_row}'] = row[col_name]
                
    #-------------------Results dispatched ---------------------results dispatch.sql
    Dispatch_mappings = {
        
           'Plasma_RECEIVED_Results_dispatched_by_lab': 'AE',
           'DBS_RECEIVED_Results_dispatched_by_lab' : 'AF',
           'Plasma_hub_dispatched_by_lab' : 'AG',
           'DBS_hub_dispatched_by_lab': 'AH'
        
              }

    # Iterate through each row in the DataFrame
    for index, row in df_dispatch.iterrows():
        # Find the corresponding row in the Excel sheet
        disp_lab_name = row['lab']

        disp_excel_row = None
        for row_num in range(start_row, start_row + 14):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == disp_lab_name:
                disp_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if disp_excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in Dispatch_mappings.items():          
                sheet[f'{excel_col}{disp_excel_row}'] = row[col_name]



    #---------referal reasons ---------referrel reasons.sql

    Ref_Reasons_mappings = {
        
           'Plasma_REFERRED_Reagent_Stockout' : 'AZ', 
           'DBS_REFERRED_Reagent_Stockout': 'BA',
           'Plasma_REFERRED_Instrument_Failure': 'BB',
           'DBS_REFERRED_Instrument_Failure' : 'BC',
           'Plasma_REFERRED_Insuff_Instrument_Capacity' : 'BD',
           'DBS_REFERRED_Insuff_Instrument_Capacity': 'BE',
           'Plasma_REFERRED_Insuff_HR_Capacity': 'BF',
           'DBS_REFERRED_Insuff_HR_Capacity':'BG',
           'Plasma_REFERRED_Other_rexns':'BH' ,
           'DBS_REFERRED_Other_rexns':'BI' 

        
              }


    # Iterate through each row in the DataFrame
    for index, row in df_ref_reasons.iterrows():
        # Find the corresponding row in the Excel sheet
        ref_rxn_lab_name = row['lab']

        ref_rxn_excel_row = None
        for row_num in range(start_row, start_row + 14):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == ref_rxn_lab_name:
                ref_rxn_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if ref_rxn_excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in Ref_Reasons_mappings.items():          
                sheet[f'{excel_col}{ref_rxn_excel_row}'] = row[col_name]



    #--------EID Populations 

    EID_mappings = {

        'Carryover_Samples_in_the_lab': 'E',
        'Samples_in_backlog_Intra_lab_TAT_7_days': 'G',
        'Total_samples_received': 'I',  
        'Num_Rejected_Samples': 'K',
        'LIMS_logged_during_week_of_arrival': 'O',
        'LIMS_hub_logged_prior_to_arrival': 'Q',

        'REJECTED_too_old_to_test' :'AK',
        'REJECTED_Quality_issue' : 'AM',
        'REJECTED_Quantity_insuff' : 'AO',
       'REJECTED_Quanti_Quali_intransit_compromised' : 'AQ',
       'REJECTED_Patient_SampleINFO' : 'AS',
       'REJECTED_Missing_requestForm' : 'AU',
        'REJECTED_Sample_Missing' :'AW',
        'REJECTED_other_reasons' : 'BY',
       'combined_comments' : 'BJ'
    }


    EID_start_row = 29
    # Iterate through each row in the DataFrame
    for index, row in df_eid_rec.iterrows():
        # Find the corresponding row in the Excel sheet
        eid_lab_name = row['Lab']

        eid_excel_row = None
        for row_num in range(EID_start_row, EID_start_row + 6):  # Assuming there are 6 labs
            if sheet[f'C{row_num}'].value == eid_lab_name:
                eid_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if eid_excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in EID_mappings.items():          
                sheet[f'{excel_col}{eid_excel_row}'] = row[col_name]

                
                
    # ---------VL reasons - for failure in sample run----------Sample Run Reasons VL.sql

    Run_rexns_mappings = {
        
           'Roche_Plasma_sample_handling_error' : 'D',  
           'Roche_Plasma_Quality_issue' : 'E',
           'Roche_Plasma_QC_Failure' : 'F',
           'Roche_Plasma_Power_Failure' : 'G', 
           'Roche_Plasma_Mechanical_Failure' : 'H',   
           'Roche_Plasma_Handling_Error':'I',
            'Roche_Plasma_quality_quantity_issues':'J',
            'Roche_Plasma_OTHER_Failure' : 'K',
            
           'Roche_DBS_sample_handling_error':'L', 
           'Roche_DBS_Quality_issue' : 'M', 
           'Roche_DBS_QC_Failure': 'N',
           'Roche_DBS_Power_Failure': 'O', 
           'Roche_DBS_Mechanical_Failure': 'P',         
           'Roche_DBS_Handling_Error':'Q',
           'Roche_DBS_quality_quantity_issues':'R',
           'Roche_DBS_OTHER_Failure': 'S',
           
           'Cobas_Plasma_sample_handling_error':'T',         
           'Cobas_Plasma_Quality_issue':'U',
           'Cobas_Plasma_QC_Failure':'V',
           'Cobas_Plasma_Power_Failure':'W',
           'Cobas_Plasma_Mechanical_Failure':'X',  
           'Cobas_Plasma_Handling_Error':'Y',      
           'Cobas_Plasma_quality_quantity_issues':'Z',
           'Cobas_Plasma_OTHER_Failure':'AA',
            
           'Cobas_DBS_sample_handling_error':'AB',           
           'Cobas_DBS_Quality_issue' :'AC', 
           'Cobas_DBS_QC_Failure': 'AD',
           'Cobas_DBS_Power_Failure': 'AE',
           'Cobas_DBS_Mechanical_Failure': 'AF',  
           'Cobas_DBS_Handling_Error': 'AG', 
           'Cobas_DBS_quality_quantity_issues':'AH',   
           'Cobas_DBS_OTHER_Failure': 'AI',
           
           
           'Abbott_Plasma_sample_handling_error':'AJ',
           'Abbott_Plasma_Quality_issue': 'AK', 
           'Abbott_Plasma_QC_Failure': 'AL',
           'Abbott_Plasma_Power_Failure':'AM', 
           'Abbott_Plasma_Mechanical_Failure':'AN', 
           'Abbott_Plasma_Handling_Error': 'AO',
           'Abbott_Plasma_quality_quantity_issues':'AP',
           'Abbott_Plasma_OTHER_Failure':'AQ',
            
           'Abbott_DBS_sample_handling_error':'AR',           
           'Abbott_DBS_Quality_issue':'AS', 
           'Abbott_DBS_QC_Failure':'AT',
           'Abbott_DBS_Power_Failure':'AU',
           'Abbott_DBS_Mechanical_Failure':'AV', 
           'Abbott_DBS_Handling_Error':'AW',        
           'Abbott_DBS_quality_quantity_issues':'AX',
           'Abbott_DBS_OTHER_Failure':'AY',
            
           'Hologic_Plasma_sample_handling_error':'AZ',          
           'Hologic_Plasma_Quality_issue': 'BA', 
           'Hologic_Plasma_QC_Failure': 'BB',
           'Hologic_Plasma_Power_Failure': 'BC', 
           'Hologic_Plasma_Mechanical_Failure': 'BD',
            'Hologic_Plasma_Handling_Error' : 'BE',  
           'Hologic_Plasma_quality_quantity_issues':'BF',   
           'Hologic_Plasma_OTHER_Failure': 'BG',
           
           
           'Hologic_DBS_sample_handling_error':'BH',          
           'Hologic_DBS_Quality_issue':'BI', 
           'Hologic_DBS_QC_Failure':'BJ',
           'Hologic_DBS_Power_Failure':'BK',
           'Hologic_DBS_Mechanical_Failure':'BL', 
           'Hologic_DBS_Handling_Error':'BM',
           'Hologic_DBS_quality_quantity_issues':'BN' ,     
           'Hologic_DBS_OTHER_Failure':'BO' 
          
           
              
    }


    VL_runRxns_start_row = 73
    # Iterate through each row in the DataFrame
    for index, row in df_vl_run_reasons.iterrows():
        # Find the corresponding row in the Excel sheet
        run_rxn_lab_name = row['name_of_lab']

        runxc_excel_row = None
        for row_num in range(VL_runRxns_start_row, VL_runRxns_start_row + 14):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == run_rxn_lab_name:
                runxc_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if runxc_excel_row:
            
            # Write the data to the corresponding cells
            for col_name, excel_col in Run_rexns_mappings.items():          
                sheet[f'{excel_col}{runxc_excel_row}'] = row[col_name]
                

     #------------------------EID Sample Run Reasons for failure----------------------Sample Run Reasons EID.sql    

    EIDrun_rexns_mappings = {
        
              
           'Roche_EID_sample_handling_error':'L', 
           'Roche_EID_Quality_issue' : 'M', 
           'Roche_EID_QC_Failure': 'N',
           'Roche_EID_Power_Failure': 'O', 
           'Roche_EID_Mechanical_Failure': 'P',         
           'Roche_EID_Handling_Error':'Q',
           'Roche_EID_quality_quantity_issues':'R',
           'Roche_EID_OTHER_Failure': 'S',
           
      
            
           'Cobas_EID_sample_handling_error':'AB',           
           'Cobas_EID_Quality_issue' :'AC', 
           'Cobas_EID_QC_Failure': 'AD',
           'Cobas_EID_Power_Failure': 'AE',
           'Cobas_EID_Mechanical_Failure': 'AF',  
           'Cobas_EID_Handling_Error': 'AG', 
           'Cobas_EID_quality_quantity_issues':'AH',   
           'Cobas_EID_OTHER_Failure': 'AI',
           
           

            
           'Abbott_EID_sample_handling_error':'AR',           
           'Abbott_EID_Quality_issue':'AS', 
           'Abbott_EID_QC_Failure':'AT',
           'Abbott_EID_Power_Failure':'AU',
           'Abbott_EID_Mechanical_Failure':'AV', 
           'Abbott_EID_Handling_Error':'AW',        
           'Abbott_EID_quality_quantity_issues':'AX',
           'Abbott_EID_OTHER_Failure':'AY',
            
 
           
           'Hologic_EID_sample_handling_error':'BH',          
           'Hologic_EID_Quality_issue':'BI', 
           'Hologic_EID_QC_Failure':'BJ',
           'Hologic_EID_Power_Failure':'BK',
           'Hologic_EID_Mechanical_Failure':'BL', 
           'Hologic_EID_Handling_Error':'BM',
           'Hologic_EID_quality_quantity_issues':'BN' ,     
           'Hologic_EID_OTHER_Failure':'BO' 
          

              
    }


    EID_runRxns_start_row = 89
    # Iterate through each row in the DataFrame
    for index, row in df_eid_run_reasons.iterrows():
        # Find the corresponding row in the Excel sheet
        EID_rxn_lab_name = row['name_of_lab']

        EIDxc_excel_row = None
        for row_num in range(EID_runRxns_start_row, EID_runRxns_start_row + 6):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == EID_rxn_lab_name:
                EIDxc_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if EIDxc_excel_row:        
            # Write the data to the corresponding cells
            for col_name, excel_col in EIDrun_rexns_mappings.items():          
                sheet[f'{excel_col}{EIDxc_excel_row}'] = row[col_name]
                
    #----------------------------SAMPLE RUN.sql

    samplerun_mapping={    
        
           'Roche_Plasma_Sample_Run' : 'D',
           'Roche_Plasma_Fail_elig_repeat': 'E', 
           'Roche_Plasma_Fail_NOT_elig_repeat' : 'F',
           'Roche_Plasma_retest_run' :'G', 
            'Roche_Plasma_Failed_tests_after_final' : 'H',
        
           'Roche_DBS_Sample_Run' : 'I',
           'Roche_DBS_Fail_elig_repeat' : 'J',
           'Roche_DBS_Fail_NOT_elig_repeat' : 'K',
           'Roche_DBS_retest_run' : 'L',
           'Roche_DBS_Failed_tests_after_final': 'M',
           
           'Cobas_Plasma_Sample_Run' : 'N',
           'Cobas_Plasma_Fail_elig_repeat' : 'O',
           'Cobas_Plasma_Fail_NOT_elig_repeat' : 'P',
           'Cobas_Plasma_retest_run' : 'Q', 
           'Cobas_Plasma_Failed_tests_after_final' :'R',
        
           'Cobas_DBS_Sample_Run' : 'S', 
           'Cobas_DBS_Fail_elig_repeat' : 'T',
           'Cobas_DBS_Fail_NOT_elig_repeat' : 'U',
           'Cobas_DBS_retest_run' : 'V',
           'Cobas_DBS_Failed_tests_after_final' : 'W',
           
           'Abbott_Plasma_Sample_Run' : 'X',
           'Abbott_Plasma_Fail_elig_repeat' : 'Y', 
           'Abbott_Plasma_Fail_NOT_elig_repeat': 'Z',
           'Abbott_Plasma_retest_run' : 'AA', 
           'Abbott_Plasma_Failed_tests_after_final': 'AB',
        
           'Abbott_DBS_Sample_Run' : 'AC', 
           'Abbott_DBS_Fail_elig_repeat' : 'AD',
           'Abbott_DBS_Fail_NOT_elig_repeat' : 'AE', 
           'Abbott_DBS_retest_run' : 'AF',
           'Abbott_DBS_Failed_tests_after_final' : 'AG',
           
           'Hologic_Plasma_Sample_Run' : 'AH',
           'Hologic_Plasma_Fail_elig_repeat' : 'AI',
           'Hologic_Plasma_Fail_NOT_elig_repeat' : 'AJ', 
           'Hologic_Plasma_retest_run' : 'AK',
           'Hologic_Plasma_Failed_tests_after_final' : 'AL',
           
           'Hologic_DBS_Sample_Run' : 'AM',
           'Hologic_DBS_Fail_elig_repeat' : 'AN', 
           'Hologic_DBS_Fail_NOT_elig_repeat' : 'AO',
           'Hologic_DBS_retest_run' : 'AP',
           'Hologic_DBS_Failed_tests_after_final' : 'AQ'
         
    }

    # Starting row in the Excel sheet
    sampleRUN_start_row = 45

    # Iterate through each row in the DataFrame
    for index, row in df_VL_run.iterrows():
        # Find the corresponding row in the Excel sheet
        runlab_name = row['name_of_lab']

        srunexcel_row = None
        for row_num in range(sampleRUN_start_row, sampleRUN_start_row + 14):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == runlab_name:
                srunexcel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if srunexcel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in samplerun_mapping.items():          
                sheet[f'{excel_col}{srunexcel_row}'] = row[col_name]
                
                
                
    #------------------EID run ------- Sample RUN EID.sql

    EID_samplerun_mapping={    

           'Roche_EID_Sample_Run' : 'I',
           'Roche_EID_Fail_elig_repeat' : 'J',
           'Roche_EID_Fail_NOT_elig_repeat' : 'K',
           'Roche_EID_retest_run' : 'L',
           'Roche_EID_Failed_tests_after_final': 'M',

        
           'Cobas_EID_Sample_Run' : 'S', 
           'Cobas_EID_Fail_elig_repeat' : 'T',
           'Cobas_EID_Fail_NOT_elig_repeat' : 'U',
           'Cobas_EID_retest_run' : 'V',
           'Cobas_EID_Failed_tests_after_final' : 'W',

        
           'Abbott_EID_Sample_Run' : 'AC', 
           'Abbott_EID_Fail_elig_repeat' : 'AD',
           'Abbott_EID_Fail_NOT_elig_repeat' : 'AE', 
           'Abbott_EID_retest_run' : 'AF',
           'Abbott_EID_Failed_tests_after_final' : 'AG',
           

           'Hologic_EID_Sample_Run' : 'AM',
           'Hologic_EID_Fail_elig_repeat' : 'AN', 
           'Hologic_EID_Fail_NOT_elig_repeat' : 'AO',
           'Hologic_EID_retest_run' : 'AP',
           'Hologic_EID_Failed_tests_after_final' : 'AQ'
         
    }

    # Starting row in the Excel sheet
    EID_sampleRUN_start_row = 62

    # Iterate through each row in the DataFrame
    for index, row in df_EID_run.iterrows():
        # Find the corresponding row in the Excel sheet
        EID_runlab_name = row['name_of_lab']

        EID_srunexcel_row = None
        for row_num in range(EID_sampleRUN_start_row, EID_sampleRUN_start_row + 6):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == EID_runlab_name:
                EID_srunexcel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if EID_srunexcel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in EID_samplerun_mapping.items():          
                sheet[f'{excel_col}{EID_srunexcel_row}'] = row[col_name]

    #--------- run comments
    Comments_samplerun_mapping={    
       'Comments': 'D'
          
    }

    # Starting row in the Excel sheet
    commentssampleRUN_start_row = 100

    # Iterate through each row in the DataFrame
    for index, row in df_VL_run.iterrows():
        # Find the corresponding row in the Excel sheet
        comm_runlab_name = row['name_of_lab']

        comm_srunexcel_row = None
        for row_num in range(commentssampleRUN_start_row, commentssampleRUN_start_row + 13):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == comm_runlab_name:
                comm_srunexcel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if comm_srunexcel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in Comments_samplerun_mapping.items():          
                sheet[f'{excel_col}{comm_srunexcel_row}'] = row[col_name]


    #---------------------------downtime -------------
    downtime_mapping={  
           'Roche_1_Downtime_Power_Outage': 'D',
           'Roche_1_Downtime_Mechanical_Failure':'E',
           'Roche_1_Downtime_Reagent_Stockout_Expiry':'F',
           'Roche_1_Downtime_Controls_Stockout_Expiry':'G',
           'Roche_1_Downtime_Controls_Failure':'H',
           'Roche_1_Downtime_Staff_Unavailability':'I',
           'Roche_1_Downtime_coz_other_reasons':'J', 
        
           'Roche_2_Downtime_Power_Outage':'K',
           'Roche_2_Downtime_Mechanical_Failure':'L',
           'Roche_2_Downtime_Reagent_Stockout_Expiry':'M',
           'Roche_2_Downtime_Controls_Stockout_Expiry':'N',
           'Roche_2_Downtime_Controls_Failure':'O',
           'Roche_2_Downtime_Staff_Unavailability':'P',
           'Roche_2_Downtime_coz_other_reasons':'Q',
        
           'Cobas_Downtime_Power_Outage':'R',
           'Cobas_Downtime_Mechanical_Failure':'S',
           'Cobas_Downtime_Reagent_Stockout_Expiry':'T',
           'Cobas_Downtime_Controls_Stockout_Expiry':'U',
           'Cobas_Downtime_Controls_Failure':'V',
           'Cobas_Downtime_Staff_Unavailability':'W',
           'Cobas_Downtime_coz_other_reasons':'X', 
        
           'Abbott_1_Downtime_Power_Outage':'Y',
           'Abbott_1_Downtime_Mechanical_Failure':'Z',
           'Abbott_1_Downtime_Reagent_Stockout_Expiry':'AA',
           'Abbott_1_Downtime_Controls_Stockout_Expiry':'AB',
           'Abbott_1_Downtime_Controls_Failure':'AC',
           'Abbott_1_Downtime_Staff_Unavailability':'AD',
           'Abbott_1_Downtime_coz_other_reasons':'AE', 
        
           'Abbott_2_Downtime_Power_Outage':'AF',
           'Abbott_2_Downtime_Mechanical_Failure':'AG',
           'Abbott_2_Downtime_Reagent_Stockout_Expiry':'AH',
           'Abbott_2_Downtime_Controls_Stockout_Expiry':'AI',
           'Abbott_2_Downtime_Controls_Failure':'AJ',
           'Abbott_2_Downtime_Staff_Unavailability':'AK',
           'Abbott_2_Downtime_coz_other_reasons':'AL',
        
           'Hologic_Downtime_Power_Outage':'AM',
           'Hologic_Downtime_Mechanical_Failure':'AN',
           'Hologic_Downtime_Reagent_Stockout_Expiry':'AO',
           'Hologic_Downtime_Controls_Stockout_Expiry':'AP',
           'Hologic_Downtime_Controls_Failure':'AQ',
           'Hologic_Downtime_Staff_Unavailability':'AR',
           'Hologic_Downtime_coz_other_reasons':'AS',
           'Comments':'AT'
       
          
    }
    Downtime_start_row = 122
    # Iterate through each row in the DataFrame
    for index, row in df_downtime.iterrows():
        # Find the corresponding row in the Excel sheet
        downtime_lab_name = row['name_of_lab']

        downtime_excel_row = None
        for row_num in range(Downtime_start_row, Downtime_start_row + 13):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == downtime_lab_name:
                downtime_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if downtime_excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in downtime_mapping.items():          
                sheet[f'{excel_col}{downtime_excel_row}'] = row[col_name]
                
    #----------------------------Power Outage ---------------------------------------------------------->
    outage_mapping={  
           'Hours_with_no_electricity':'D', 
           'Hours_generator_was_on' :'E',
           'Fuel_ltrs_added_to_generator' :'F', 
           'Hrs_Machine_idle_coz_PowerCuts' :'G',
           'Total_Tests_done_using_generator': 'H', 
           'Comments':'I'      
         
          
    }
    outage_start_row = 140
    # Iterate through each row in the DataFrame
    for index, row in df_power_outage.iterrows():
        # Find the corresponding row in the Excel sheet
        outage_lab_name = row['Laboratory']

        outage_excel_row = None
        for row_num in range(outage_start_row, outage_start_row + 13):  # Assuming there are 14 labs
            if sheet[f'C{row_num}'].value == outage_lab_name:
                outage_excel_row = row_num
              
                break

        # If the lab name is found in the Excel sheet
        if outage_excel_row:
            # Write the data to the corresponding cells
            for col_name, excel_col in outage_mapping.items():          
                sheet[f'{excel_col}{outage_excel_row}'] = row[col_name]
    


    # Save the workbook
    workbook.save(file_path)






# View for downloading the Excel file
def download_excel(request):
    file_path = request.session.get('excel_file_path')

    if file_path and os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = FileResponse(fh, as_attachment=True)
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
    else:
        return HttpResponseNotFound()  # Return a 404 Not Found response

